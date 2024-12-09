#!/usr/bin/python3
# Clean up Banweb "new" classlist XLSX export,
# generating student-emails.txt and "new format" students.csv.
# Bart Massey 2021

import argparse, csv, sys, warnings
from collections import namedtuple

from openpyxl import load_workbook

import tiles

ap = argparse.ArgumentParser()
ap.add_argument("-d", "--disc", type=float, default=0.7)
ap.add_argument("sheet")
args = ap.parse_args()

sheet = args.sheet
disc = args.disc

Record = namedtuple(
    'Record',
    ['year', 'term', 'cid', 'title', 'instructor', 'tiles'],
)

quarters = ["Wi", "Sp", "Su", "Fa"]

records = []
with warnings.catch_warnings(record=True):
    warnings.simplefilter("always")
    wb = load_workbook(filename=sheet)
    ws = wb.active
    v = ws.cell(column=1, row=1).value
    if v != "Subject":
        print(f"{section_file}: sheet header error", file=sys.stderr)
        exit(1)

    headers = { c.value : c.column - 1 for (c,) in ws.iter_cols(
        min_row=1,
        max_row=1,
    ) }
    term_header = headers["Term"]
    crn_header = headers["CRN"]
    status_header = headers["Status"]
    title_header = headers["Title"]
    instructor_last_header = headers["Instructor_Last_Name"]
    instructor_first_header = headers["Instructor_First_Name"]

    for row in ws.iter_rows(min_row=2, values_only=True):
        yearterm = row[term_header]
        year = int(yearterm[:4])
        term = int(yearterm[4:])

        crn = row[crn_header]
        cid = crn + "-" + str(year)
        status = row[status_header]
        if status == "Cancelled":
            continue

        title = row[title_header]
        if not title.startswith("TOP:"):
            continue
        title = title[4:].strip()

        instructor_last = row[instructor_last_header]
        # XXX Weird special case for David Whitlock. Idk.
        whitlock = " Comp. Sci"
        if instructor_last and instructor_last.endswith(whitlock):
            instructor_last = instructor_last[:-len(whitlock)]
        instructor_first = row[instructor_first_header]
        if instructor_first and instructor_last:
            instructor = instructor_last + ", " + instructor_first
        else:
            if instructor_last:
                instructor = instructor_last
            else:
                instructor = ""

        ts = tiles.tileset(title)

        r = Record(year, term, cid, title, instructor, ts)
        records.append(r)

records.sort(key = lambda r: (r.year, r.term), reverse=True)
nrecords = len(records)

simcount = 0
simsum = 0
smatrix = [[None] * nrecords for _ in range(nrecords)]
for i, r1 in enumerate(records):
    for j, r2 in enumerate(records):
        if i == j:
            smatrix[i][j] = 1
            continue
        simcount += 1
        s = tiles.sim(r1.tiles, r2.tiles)
        simsum += s
        smatrix[i][j] = s
simavg = simsum / simcount
threshold = simavg + (1 - simavg) * disc

clusters = []
placed = set()
for i, r1 in enumerate(records):
    if i in placed:
        continue
    cur_cluster = [i]
    placed |= {i}
    for j in range(i + 1, nrecords):
        if j in placed:
            continue
        r2 = records[j]
        if smatrix[i][j] >= threshold:
            cur_cluster.append(j)
            placed |= {j}
    clusters.append(cur_cluster)

with open("report.txt", "w") as f:
    for c in clusters:
        r0 = records[c[0]]
        print(r0.cid, r0.title, file=f)
        for cx in c[1:]:
            r = records[cx]
            print("   ", r.cid, r.title, file=f)

firsts = []
latests = []
for c in clusters:
    firsts.append(records[c[0]])
    latests.append(records[c[-1]])
latests.sort(key = lambda r: (r.year, r.term), reverse = True)

def write_csv(fn, rs):
    with open(fn, "w") as f:
        w = csv.writer(f)
        for r in rs:
            quarter = quarters[r.term - 1]
            w.writerow([r.title, r.instructor, r.year, quarter])

write_csv("latest.csv", latests)
write_csv("first.csv", latests)
